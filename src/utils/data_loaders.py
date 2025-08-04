"""Загрузчики данных из различных источников"""
import pandas as pd
import json
import openpyxl
import streamlit as st
from elasticsearch import Elasticsearch
from typing import Tuple, Optional
from config import config


class ElasticsearchLoader:
    """Загрузчик данных из Elasticsearch, включая инициализацию клиента и извлечение таблиц."""

    def __init__(self):
        self._client = None

    def get_client(self) -> Optional[Elasticsearch]:
        """Создает и возвращает подключение к Elasticsearch. Показывает предупреждение при неудаче."""
        if self._client is None:
            try:
                self._client = Elasticsearch(hosts=config.ES_HOST, verify_certs=False)
                if not self._client.ping():
                    st.warning("⚠️ Не удается подключиться к Elasticsearch.")
                    self._client = None
            except Exception as e:
                st.warning(f"⚠️ Ошибка подключения к Elasticsearch: {str(e)}")
                self._client = None
        return self._client

    def extract_table(self, report_id: str) -> Tuple[Optional[pd.DataFrame], str]:
        """
        Извлекает таблицу из Elasticsearch по ID документа.
        
        Аргументы:
            report_id (str): Идентификатор документа в Elasticsearch.
        
        Возвращает:
            Tuple[Optional[pd.DataFrame], str]: DataFrame с таблицей и статусное сообщение.
        """
        try:
            es = self.get_client()
            if es is None:
                return None, "Нет подключения к Elasticsearch"

            response = es.get(index=config.ES_INDEX, id=report_id)
            if response['found']:
                table_json = response['_source']['table_data_json']
                df = pd.DataFrame(json.loads(table_json))
                return df, "Система готова к работе!"
            else:
                return None, f"Отчет с ID {report_id} не найден в Elasticsearch"
        except Exception as e:
            return None, f"Ошибка при извлечении данных из Elasticsearch: {str(e)}"


class ExcelLoader:
    """Загрузчик и парсер Excel-файлов замеров, газов и шахматок для анализа."""

    def parse_drilling_excel(self, file_path: str) -> pd.DataFrame:
        """
        Парсит Excel-файл шахматки проб (бурения).
        
        Аргументы:
            file_path (str): Путь к Excel-файлу.
        
        Возвращает:
            pd.DataFrame: Пустой DataFrame (заглушка, логика не реализована).
        """
        # Здесь должна быть логика парсинга файла шахматки
        # Но меня  попросили именно эту часть скрыть
        return pd.DataFrame()

    def parse_measurement_excel(self, file_path: str) -> pd.DataFrame:
        """
        Парсит Excel-файл замерной добычи.
        
        Формирует DataFrame из диапазонов и ячеек, корректно обрабатывая объединенные ячейки,
        исключает невалидные строки и очищает от итогов.
        
        Аргументы:
            file_path (str): Путь к Excel-файлу.
        
        Возвращает:
            pd.DataFrame: Чистый и структурированный DataFrame с измерениями.
        """
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            worksheet = workbook['Лист1']

            merged_cells = worksheet.merged_cells.ranges

            df_raw = pd.read_excel(file_path, sheet_name='Лист1', header=None)

            header_row_1 = 13 
            header_row_2 = 14 
            data_start_row = 15  
            data_end_row = 499  
            data_start_col = 16 
            data_end_col = 52 

            excluded_rows_excel = [17, 27, 28, 36, 59, 84, 92, 93, 96, 136, 138, 139, 144, 146, 147, 149,
                                   156, 165, 166, 170, 171, 173, 174, 186, 198, 215, 219, 220, 223, 224, 226,
                                   248, 249, 254, 255, 333, 338, 341, 346, 347, 349, 350, 352, 356, 371, 372, 374,
                                   375, 377, 378, 382, 386, 388, 389, 400, 402, 403, 405, 408, 414, 419, 421,
                                   423, 427, 429, 430, 438, 443, 444, 465, 467, 468, 492, 493]

            def get_merged_cell_value(row, col):
                for merged_range in merged_cells:
                    min_col, min_row, max_col, max_row = merged_range.min_col, merged_range.min_row, merged_range.max_col, merged_range.max_row
                    if (min_row <= row + 1 <= max_row) and (min_col <= col + 1 <= max_col):
                        return df_raw.iloc[min_row - 1, min_col - 1]
                return df_raw.iloc[row, col]

            headers = []
            for col in range(data_start_col, data_end_col + 1):
                header_1 = get_merged_cell_value(header_row_1, col)
                header_2 = get_merged_cell_value(header_row_2, col)

                header_1 = str(header_1) if pd.notna(header_1) else ""
                header_2 = str(header_2) if pd.notna(header_2) else ""

                if header_1 and header_2:
                    if header_1.strip() == header_2.strip():
                        combined_header = header_1
                    else:
                        combined_header = f"{header_1}, {header_2}"
                elif header_1:
                    combined_header = header_1
                elif header_2:
                    combined_header = header_2
                else:
                    excel_col = openpyxl.utils.get_column_letter(col + 1)
                    combined_header = f"Столбец_{excel_col}"

                combined_header = combined_header.replace('\n', ' ').strip()
                headers.append(combined_header)

            df = df_raw.iloc[data_start_row:data_end_row + 1, data_start_col:data_end_col + 1].copy()

            excluded_rows_indices = [row - 1 for row in excluded_rows_excel]
            original_row_numbers = [idx + data_start_row + 1 for idx in range(len(df))]

            rows_to_exclude_mask = []
            for i, excel_row_num in enumerate(original_row_numbers):
                should_exclude = excel_row_num in excluded_rows_excel
                rows_to_exclude_mask.append(should_exclude)

            include_mask = [not exclude for exclude in rows_to_exclude_mask]
            df = df[include_mask].copy()

            df = df.dropna(how='all')

            if len(df.columns) == len(headers):
                df.columns = headers
            else:
                if len(headers) < len(df.columns):
                    headers.extend([f"Столбец_{i + 1}" for i in range(len(headers), len(df.columns))])
                df.columns = headers[:len(df.columns)]

            df = df.reset_index(drop=True)

            mestorozhdenie_col = 'Месторождение'

            if mestorozhdenie_col in df.columns:
                is_summary_row = df[mestorozhdenie_col].astype(str).str.contains(
                    'итог|всего|цднг|сумм|total', case=False, na=False
                )

                well_col = '№ скв'
                if well_col in df.columns:
                    is_empty_well = df[well_col].isna() | (df[well_col].astype(str).str.strip() == '')

                    additional_summary = is_empty_well & df[mestorozhdenie_col].astype(str).str.contains(
                        'цднг|сумм|итог', case=False, na=False
                    )

                    is_summary_row = is_summary_row | additional_summary

                if is_summary_row.any():
                    df = df[~is_summary_row].copy()

            df = df.reset_index(drop=True)

            numeric_columns = []
            for col in df.columns:
                if any(keyword in col.lower() for keyword in
                       ['qж', 'qн', '%', 'т/сут', 'м3/сут', 'дебит', 'ку', 'период']):
                    numeric_columns.append(col)

            for col in numeric_columns:
                df[col] = pd.to_numeric(df[col], errors='coerce')

            return df

        except Exception as e:
            raise Exception(f"Ошибка при загрузке файла замерной добычи: {str(e)}")

    def parse_gas_excel(self, file_path: str) -> pd.DataFrame:
        """
        Парсит Excel-файл утилизации газа.
        
        Обрабатывает многострочные заголовки с объединением ячеек, отсекает итоговые строки и формирует финальный DataFrame.
        
        Аргументы:
            file_path (str): Путь к Excel-файлу.
        
        Возвращает:
            pd.DataFrame: Структурированный DataFrame по утилизации газа.
        
        """
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            worksheet = workbook['Лист1']

            merged_cells = worksheet.merged_cells.ranges

            df_raw = pd.read_excel(file_path, sheet_name='Лист1', header=None)

            header_start_row = 18 
            header_end_row = 24  
            data_start_row = 26
            data_end_row = 43 

            def get_merged_cell_value(row, col):
                for merged_range in merged_cells:
                    min_col, min_row, max_col, max_row = merged_range.min_col, merged_range.min_row, merged_range.max_col, merged_range.max_row
                    if (min_row <= row + 1 <= max_row) and (min_col <= col + 1 <= max_col):
                        return df_raw.iloc[min_row - 1, min_col - 1]
                return df_raw.iloc[row, col]

            headers = []
            for col in range(8, 46):  
                header_parts = []
                for row in range(header_start_row, header_end_row + 1):
                    cell_value = get_merged_cell_value(row, col)
                    if pd.notna(cell_value):
                        header_parts.append(str(cell_value))

                filtered_parts = [part for part in header_parts if not part.startswith("Отчет о выполнении утилизации")]

                seen = set()
                unique_parts = []
                for part in filtered_parts:
                    if part not in seen:
                        seen.add(part)
                        unique_parts.append(part)

                if unique_parts:
                    header_name = " ".join(unique_parts)
                    headers.append(header_name)
                else:
                    headers.append(f"Столбец_{col + 1}")

            df = df_raw.iloc[data_start_row:data_end_row, 8:46].copy()

            df = df.dropna(how='all')

            if len(df.columns) == len(headers):
                df.columns = headers
            else:
                if len(headers) < len(df.columns):
                    headers.extend([f"Столбец_{i + 1}" for i in range(len(headers), len(df.columns))])
                df.columns = headers[:len(df.columns)]

            first_col = df.columns[0]

            is_summary_row = df[first_col].astype(str).str.contains('итог|всего|квартал|итого|месяц', case=False,
                                                                    na=False)
            quarter_pattern = r'^(I|II|III|IV)\s+квартал|^(I|II|III|IV)$|год'
            is_quarter_row = df[first_col].astype(str).str.match(quarter_pattern, case=False, na=False)

            rows_to_drop = is_summary_row | is_quarter_row

            if rows_to_drop.any():
                df = df[~rows_to_drop].copy()

            df = df.reset_index(drop=True)

            return df

        except Exception as e:
            raise Exception(f"Ошибка при загрузке файла утилизации газа: {str(e)}")